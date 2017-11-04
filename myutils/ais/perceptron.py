import MySQLdb
import pandas as pd
import numpy as np
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill, Color

import matplotlib.pyplot as plt
import itertools

from sklearn.metrics import confusion_matrix
from sklearn.metrics import accuracy_score
from sklearn.metrics import f1_score
from sklearn.ensemble import RandomForestClassifier


###############   CONST   #############
PCA_DIM = 10
LEARNING_START_DATE = '2017-03-01'
LEARNING_END_DATE = '2017-07-14'
#######################################



# PCA
from sklearn.decomposition import PCA
def do_pca(dataset,n):
    pca = PCA(n_components=n)
    pca.fit(dataset)
    print(pca.explained_variance_ratio_)
    x_train_transformed = pca.fit_transform(dataset)
    #print(x_train_transformed)
    return x_train_transformed


def plot_confusion_matrix(cm, classes,normalize=True,title='Confusion Matrix',cmap=plt.cm.Blues):
    plt.imshow(cm, interpolation='nearest',cmap=cmap)
    plt.title(title)
    plt.colorbar()
    tick_marks = np.arange(len(classes))
    plt.xticks(tick_marks, classes, rotation=45)
    plt.yticks(tick_marks, classes)
    
    if normalize :
        cm = cm.astype('float') / cm.sum(axis = 1)[:,np.newaxis]
    else:
        print("confusion Matrix, without normalization")
    
    print(cm)
    
    thresh = cm.max()/2.
    for i, j in itertools.product(range(cm.shape[0]),range(cm.shape[1])):
        plt.text(j,i,cm[i,j],
                horizontalalignment="center",
                color="white" if cm[i,j] > thresh else "black")
    
    plt.tight_layout()
    plt.ylabel('True label')
    plt.xlabel('predicted label')
    plt.show()
    
def do_confusion(test,pred):

    cnf_matrix = confusion_matrix(test,pred)
    print(cnf_matrix.shape)
    # bl을 bl로 맞춘 정확도
    #accuracy = cnf_matrix[1][1]/(cnf_matrix[1][0]+cnf_matrix[1][1])
    
    plot_confusion_matrix(cnf_matrix, classes=['wl','bl'])
    
    return accuracy_score(test, pred)
    #return accuracy

def do_f1(test, pred):
    return f1_score(test, pred)

from sklearn import tree

def do_decision(x_train, y_train, test):
    clf = tree.DecisionTreeClassifier()
    clf.fit(x_train, y_train)
    y_pred = clf.predict(test)
    
    return y_pred

from sklearn.svm import SVC
from sklearn.model_selection import GridSearchCV
#from sklearn.grid_search import GridSearchCV
from sklearn.metrics import classification_report
from sklearn.neural_network import MLPClassifier

def do_nn(x,y,test,layers,a):
    clf = MLPClassifier(solver='adam',alpha=a,hidden_layer_sizes=(PCA_DIM,layers),random_state=1)
    clf.fit(x,y)
    y_pred = clf.predict(test)
    return y_pred

def do_svm(x,y,test,c,g):
    print('x ==> ',x)
    print('y ==> ',y)
    print('test ==> ',test)
    clf = SVC(C=c, kernel='rbf',gamma=g) 
    print("fit => ", clf.fit(x,y))
    y_pred = clf.predict(test)
    print("y_pred = ", y_pred)
    return y_pred

def tune_svm(x,y,test):
    print('x => ', x)
    print('y => ',len(y==0), y)
    print('test => ', test)
    tuned_parameters = [{'kernel':['rbf'], 
                         'gamma':[1e-2,1e-3,1e-4,1e-5,1e-6,1e-7,1e-8,1e-9], 
                         'C' : [10,100,1000,10000,100000,1000000,10000000]}]
    
    scores = ['precision','recall']
    
    for score in scores:
        print("# Tunning hyper-parameters for %s" % score)
        print()
        clf = GridSearchCV(SVC(C=1), tuned_parameters, cv=5, scoring='%s_macro' % score)
        clf.fit(x,y)
        
        print("Best parameters set found on development set:")
        print()
        print(clf.best_params_)
        print("Grid scores on development set:")
        print()
        
        means = clf.cv_results_['mean_test_score']
        stds = clf.cv_results_['std_test_score']
        
        for mean, std, params in zip(means, stds, clf.cv_results_['params']):
            print("%0.3f (+/-%0.03f) for %r" % (mean, std *2, params))

        print()

        print("Detailed classification report:")
        print()
        print("The model is trained on the full deveopment set.")
        print("The scores are computed on the full evaluation set.")
        print()
        y_true, y_pred = y_test, clf.predict(test)
        print(classification_report(y_true, y_pred))
        print()
def do_rf(x, y, test):
    clf = RandomForestClassifier(n_estimators=20)
    clf.fit(x, y)    
    y_pred = clf.predict(test)   
    return y_pred

# db에서 학습 할 데이터 불러와서 전처리
def data_preprocessing(start_date, end_date):
    curs = conn.cursor()
    # bl
    train_ydata_s = 'select ip, bl_ibm, times, web_ok, web_rej, fw_pd, fw_db, waf_ok, waf_rej, ips_pd, ips_db, bl_dg_yn from aw_ip_cache where write_date >= %s and write_date < %s and (ip_gubun = 1 or ip_gubun = 2) and bl_ibm is not null'
    curs.execute(train_ydata_s,(start_date,end_date))
    train_yrows = curs.fetchall()    
    train_ydf = pd.DataFrame(list(train_yrows), columns=['ip', 'bl_ibm', 'times', 'web_ok', 'web_rej', 'fw_pd', 'fw_db', 'waf_ok', 'waf_rej', 'ips_pd', 'ips_db',' bl_dg_yn'])       
    train_ydf['bl_dg_yn'] = 0

    # wl : bl에 비해 wl이 지나치게 많아서 수를 줄이기 위해 수집기간을 구분해서 저장
    curs = conn.cursor()
    train_xdata_s = 'select ip, bl_ibm, times, web_ok, web_rej, fw_pd, fw_db, waf_ok, waf_rej, ips_pd, ips_db, bl_dg_yn from aw_ip_cache where modi_time >= %s and modi_time < %s and (ip_gubun = 1 or ip_gubun = 2) and bl_ibm is not null and bl_dg_yn = 0 limit 5000'
    curs.execute(train_xdata_s,('2017-05-01',end_date))
    train_xrows = curs.fetchall()    
    train_xdf = pd.DataFrame(list(train_xrows), columns=['ip', 'bl_ibm', 'times', 'web_ok', 'web_rej', 'fw_pd', 'fw_db', 'waf_ok', 'waf_rej', 'ips_pd', 'ips_db',' bl_dg_yn'])       
    train_xdf['bl_dg_yn'] = 0


    # 예측할 오늘 데이터
    test_data_s = 'select ip, bl_ibm, times, web_ok, web_rej, fw_pd, fw_db, waf_ok, waf_rej, ips_pd, ips_db, bl_dg_yn from aw_ip_cache where modi_time like %s and (ip_gubun = 1 or ip_gubun = 2) and bl_ibm is not null'
    curs.execute(test_data_s,(str(date.today())+'%',))
    test_rows = curs.fetchall()    
    test_df = pd.DataFrame(list(test_rows), columns=['ip', 'bl_ibm', 'times', 'web_ok', 'web_rej', 'fw_pd', 'fw_db', 'waf_ok', 'waf_rej', 'ips_pd', 'ips_db',' bl_dg_yn'])
    test_df['bl_dg_yn'] = 0


    
    bl = pd.read_csv('bl.csv',names=['ip'])    
    # bl.csv 에 기록된 black list를 읽어 실제 차단한 ip를 기록
    for bl_ip in bl['ip'] :
        train_ydf.ix[train_ydf['ip']==bl_ip,'bl_dg_yn'] = 1        

    print(train_ydf.head())

    print("total bl count : ",len(train_ydf.loc[train_ydf['bl_dg_yn']==1]))  
    train_y=train_ydf.ix[train_ydf['bl_dg_yn']==1]
    train_df = np.concatenate([train_y,train_xdf])
    
    #print(train_df[:,-1])
    
    #train_y_label = train_df[:,-1]
    #test_y_label = test_df['bl_dg_yn']
    test_ip = test_df['ip']

    y_train = train_df[:,-1].astype(int) 
    y_test = test_df['bl_dg_yn'].astype(int)

    x_train = train_df[:,1:-2].astype(int)
    x_test = np.asarray(test_df)
    x_test = x_test[:,1:-2].astype(int)
        
    y_pred = do_svm(x_train, y_train, x_test, 100000, 0.0001)
    
    test_y_label = y_pred
    report = np.column_stack((test_ip,test_y_label))

    bl_idx = np.where(report[:,1] == 1.0)    
    svm_res = report[bl_idx]
    svm_res = svm_res[:,0]
    print("svm res => ", svm_res)
    
    y_pred = do_decision(x_train, y_train, x_test)
    
    test_y_label = y_pred
    report = np.column_stack((test_ip,test_y_label))

    bl_idx = np.where(report[:,1] == 1.0)
    dt_res = report[bl_idx]
    dt_res = dt_res[:,0]
    print("dt res => ", dt_res)

    y_pred = do_rf(x_train, y_train, x_test)    
    test_y_label = y_pred
    report = np.column_stack((test_ip,test_y_label))

    bl_idx = np.where(report[:,1] == 1.0)
    rf_res = report[bl_idx]
    rf_res = rf_res[:,0]
    print("rf res => ", rf_res)

    y_pred = do_nn(x_train, y_train, x_test, 5, 1e-5)
    
    test_y_label = y_pred
    report = np.column_stack((test_ip,test_y_label))

    bl_idx = np.where(report[:,1] == 1.0)
    nn_res = report[bl_idx]
    nn_res = nn_res[:,0]
    
    ins1 = np.union1d(rf_res, dt_res)
    ins2 = np.union1d(ins1, svm_res)
    ins3 = np.union1d(ins2, nn_res)

    test_data_s = 'select ip, bl_ibm, times, web_ok, web_rej, fw_pd, fw_db, waf_ok, waf_rej, ips_pd, ips_db, bl_dg_yn from aw_ip_cache where modi_time like %s and (ip_gubun = 1 or ip_gubun = 2) and bl_ibm is not null and bl_dg_yn=1'
    curs.execute(test_data_s,(str(date.today())+'%',))
    test_rows = curs.fetchall()    
    test_df = pd.DataFrame(list(test_rows), columns=['ip', 'bl_ibm', 'times', 'web_ok', 'web_rej', 'fw_pd', 'fw_db', 'waf_ok', 'waf_rej', 'ips_pd', 'ips_db',' bl_dg_yn'])
    test_df['bl_dg_yn'] = 0

    

    test_data_add = 'select ip, bl_ibm, times, web_ok, web_rej, fw_pd, fw_db, waf_ok, waf_rej, ips_pd, ips_db, bl_dg_yn from aw_ip_cache where ip=%s'
    for ips in ins3 :
        #print('ips =>',ips)
        curs.execute(test_data_add,(ips,))
        add_rows = curs.fetchone()
        #print(add_rows)
        df2 = pd.DataFrame(list(add_rows),)
        #df2['bl_dg_yn'] = 0
        #print(df2)
        test_df.append(df2)
    print(test_df)

    #test_df = set(test_df) # 중뷁 제거
    test_df.drop_duplicates(['ip'],keep='last')
    test_ip = test_df['ip']
    
    y_test = test_df['bl_dg_yn'].astype(int)    
    x_test = np.asarray(test_df)
    x_test = x_test[:,1:-2].astype(int)
        

    y_pred = do_nn(x_train, y_train, x_test, 7, 1e-6)
    
    test_y_label = y_pred
    report = np.column_stack((test_ip,test_y_label))

    bl_idx = np.where(report[:,1] == 1.0)
    nn_res = report[bl_idx]
    nn_res = nn_res[:,0]
    print('nn_res =>', nn_res)


    return nn_res

def report_xl(bl_ip):
    today = date.today()
    wb = Workbook()
    ws = wb.active
    ws.title='AI'
    ws.merge_cells('A1:L1')
    ws['A1'] = str(today) + ' Threating IP Information From AI'
    ca1 = ws['A1']
    ca1.font = Font(name='맑은 고딕', size = 15, bold=True)
    ca1.alignment = Alignment(horizontal = 'center', vertical = 'center')
    box = Border(left=Side(border_style="thin", 
                   color='FF000000'),
         right=Side(border_style="thin",
                    color='FF000000'),
         top=Side(border_style="thin",
                  color='FF000000'),
         bottom=Side(border_style="thin",
                     color='FF000000'),
         diagonal=Side(border_style="thin",
                       color='FF000000'),
         diagonal_direction=0,
         outline=Side(border_style="thin",
                      color='FF000000'),
         vertical=Side(border_style="thin",
                       color='FF000000'),
         horizontal=Side(border_style="thin",
                        color='FF000000')
        )
    ca1.border = box
    ca1.fill = PatternFill(patternType='solid',fgColor=Color('FFC000'))

    #ca2 = ws['A2:L2']
    #ca2.font = Font(size = 11.5, bold=True)
    #ca2.alignment = Alignment(horizontal = 'center', vertical = 'center')
    #ca2.fill = PatternFill(patternType='solid',fgColor=Color('blue'))

    ws['A2'] = 'IP'
    ws['B2'] = '국가'
    ws['C2'] = 'IBM Score'
    ws['D2'] = '횟수'
    ws['E2'] = '지속일수'    
    ws['F2'] = 'WEB OK'
    ws['G2'] = 'FW DR'
    ws['H2'] = 'WAF OK'
    ws['I2'] = 'WAF REJ'
    ws['J2'] = 'IPS PD'
    ws['K2'] = 'IPS DB'
    ws['L2'] = 'SQL'
    ws.freeze_panes = 'A3' # 셀 고정
    curs = conn.cursor()
    #times_today
    #select ip, cc, bl_ibm,
    #i = 0
    se_ip_ca = 'select ip,cc,bl_ibm,web_ok,fw_db,waf_ok,waf_rej,ips_pd,ips_db,bl_dg_yn from aw_ip_cache where ip = %s'
    se_ti = 'select sum(times) from aw_log_full where source_ip = %s and write_date = %s'
    se_da = 'select count(distinct write_date) from aw_log_full where source_ip = %s'
    
                           
    for i,ip in enumerate(bl_ip) :
        curs.execute(se_ip_ca,(ip,))
        bl_rows = curs.fetchall()        

        curs.execute(se_ti,(ip, date.today()))        
        ti = curs.fetchone()
        
        curs.execute(se_da,(ip,))
        da = curs.fetchone()
        print('i => ',i)       
        ws['A'+str(i+3)] = bl_rows[0][0]
        ws['B'+str(i+3)] = bl_rows[0][1]
        ws['C'+str(i+3)] = bl_rows[0][2]
        ws['D'+str(i+3)] = ti[0]
        ws['E'+str(i+3)] = da[0]
        ws['F'+str(i+3)] = bl_rows[0][3]
        ws['G'+str(i+3)] = bl_rows[0][4]
        ws['H'+str(i+3)] = bl_rows[0][5]
        ws['I'+str(i+3)] = bl_rows[0][6]
        ws['J'+str(i+3)] = bl_rows[0][7]
        ws['K'+str(i+3)] = bl_rows[0][8]
        ws['L'+str(i+3)] = bl_rows[0][9]

    filename = str(date.today()) + '_report.xls'
    wb.save(filename)



    
conn = MySQLdb.connect('localhost','root','rhrnak#33','ais', charset = "utf8")
ai_bl = data_preprocessing(LEARNING_START_DATE,LEARNING_END_DATE)
ai_bl = np.array(ai_bl.tolist())
report_xl(ai_bl)


conn.close()
                     


